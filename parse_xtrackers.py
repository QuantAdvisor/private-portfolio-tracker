"""
Phase 1 – Xtrackers/DWS Holdings-Parser
Liest alle Constituent_<ISIN>.xlsx aus ETF Holdings/ und schreibt:
  - portfolio.constituent  (Einzeltitel-Stammdaten, UPSERT)
  - portfolio.etf_holding  (Gewichte je ETF + Snapshot-Datum, UPSERT)

Dateiformat:
  Zeile 1–3: Disclaimer/leer  →  skiprows=3
  Zeile 4:   Header (None, Name, ISIN, Country, Currency, Exchange,
                     Type of Security, Rating, Primary Listing,
                     Industry Classification, Weighting)
  Ab Zeile 5: Daten; Weighting = Dezimalzahl (z. B. 0.0305 = 3,05 %)
  Sheet-Name = as_of_date (z. B. '2026-06-21')

Aufruf:
    python parse_xtrackers.py
    python parse_xtrackers.py --dry-run   # zeigt was geschrieben würde, ohne DB-Zugriff
"""

import argparse
import logging
import re
from datetime import date
from pathlib import Path

import openpyxl

import db_utils

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

HOLDINGS_DIR = Path(__file__).parent / "ETF Holdings"

# ── Länder-Mapping (Deutsch → ISO-3166-1-Alpha-2) ────────────────────────────
COUNTRY_MAP: dict[str, str] = {
    "australien":                        "AU",
    "belgien":                           "BE",
    "brasilien":                         "BR",
    "china":                             "CN",
    "dänemark":                          "DK",
    "danemark":                          "DK",
    "deutschland":                       "DE",
    "finnland":                          "FI",
    "frankreich":                        "FR",
    "griechenland":                      "GR",
    "großbritannien (uk)":               "GB",
    "grossbritannien (uk)":              "GB",
    "vereinigtes königreich":            "GB",
    "vereinigtes konigreich":            "GB",
    "hongkong":                          "HK",
    "indien":                            "IN",
    "indonesien":                        "ID",
    "irland":                            "IE",
    "israel":                            "IL",
    "italien":                           "IT",
    "japan":                             "JP",
    "kanada":                            "CA",
    "katar":                             "QA",
    "korea":                             "KR",
    "luxemburg":                         "LU",
    "malaysia":                          "MY",
    "mexiko":                            "MX",
    "neuseeland":                        "NZ",
    "niederlande":                       "NL",
    "norwegen":                          "NO",
    "österreich":                        "AT",
    "osterreich":                        "AT",
    "philippinen":                       "PH",
    "polen":                             "PL",
    "portugal":                          "PT",
    "saudi-arabien":                     "SA",
    "schweden":                          "SE",
    "schweiz":                           "CH",
    "singapur":                          "SG",
    "spanien":                           "ES",
    "südafrika":                         "ZA",
    "sudafrika":                         "ZA",
    "taiwan":                            "TW",
    "thailand":                          "TH",
    "tschechische republik":             "CZ",
    "türkei":                            "TR",
    "turkei":                            "TR",
    "ungarn":                            "HU",
    "vereinigte arabische emirate":      "AE",
    "vereinigte staaten von amerika":    "US",
    "vereinigte staaten":                "US",
}


def _normalize_country(raw: str | None) -> str | None:
    if not raw:
        return None
    key = raw.lower().strip()
    # Umlaute normalisieren (für robustere Matches)
    key = key.replace("ü", "u").replace("ö", "o").replace("ä", "a").replace("ß", "ss")
    result = COUNTRY_MAP.get(key)
    if result is None:
        log.debug("Unbekanntes Land: %r – als NULL gespeichert", raw)
    return result


def _isin_from_filename(path: Path) -> str | None:
    m = re.match(r"Constituent_([A-Z0-9]{12})\.xlsx", path.name)
    return m.group(1) if m else None


# ── Datei einlesen ────────────────────────────────────────────────────────────

def parse_file(path: Path) -> tuple[str, date, list[dict], list[dict]]:
    """
    Gibt zurück: (etf_isin, as_of_date, constituent_rows, holding_rows)
    """
    etf_isin = _isin_from_filename(path)
    if not etf_isin:
        raise ValueError(f"ISIN nicht aus Dateiname ableitbar: {path.name}")

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    sheet_name = wb.sheetnames[0]

    try:
        as_of = date.fromisoformat(sheet_name)
    except ValueError:
        raise ValueError(f"Sheet-Name ist kein Datum: {sheet_name!r} in {path.name}")

    ws = wb[sheet_name]

    constituents: list[dict] = []
    holdings:     list[dict] = []
    unresolved:   int = 0

    for i, row in enumerate(ws.iter_rows(min_row=5, values_only=True)):
        # Spalten: 0=lfd, 1=Name, 2=ISIN, 3=Country, 4=Currency, 5=Exchange,
        #          6=Type, 7=Rating, 8=PrimaryListing, 9=Industry, 10=Weighting
        _, name, isin, country, currency, _, sec_type, _, _, industry, weight = row

        if not isin or not name or weight is None:
            unresolved += 1
            continue

        try:
            weight_pct = round(float(weight) * 100, 6)
        except (ValueError, TypeError):
            unresolved += 1
            continue

        if weight_pct <= 0:
            continue

        constituents.append({
            "isin":     isin.strip(),
            "name":     str(name).strip()[:200],
            "sektor":   str(industry).strip()[:100] if industry else None,
            "country":  _normalize_country(country),
            "currency": str(currency).strip()[:3] if currency else None,
        })
        holdings.append({
            "etf_isin":         etf_isin,
            "as_of_date":       as_of,
            "constituent_isin": isin.strip(),
            "weight_pct":       weight_pct,
            "source_file":      path.name,
        })

    if unresolved:
        log.debug("%s: %d Zeilen ohne ISIN/Name/Gewicht übersprungen", path.name, unresolved)

    wb.close()
    return etf_isin, as_of, constituents, holdings


# ── DB-Schreiben ──────────────────────────────────────────────────────────────

UPSERT_CONSTITUENT = """
INSERT INTO portfolio.constituent (isin, name, sektor, country, currency)
VALUES (:isin, :name, :sektor, :country, :currency)
ON CONFLICT (isin) DO UPDATE SET
    name     = EXCLUDED.name,
    sektor   = COALESCE(EXCLUDED.sektor,   portfolio.constituent.sektor),
    country  = COALESCE(EXCLUDED.country,  portfolio.constituent.country),
    currency = COALESCE(EXCLUDED.currency, portfolio.constituent.currency)
"""

UPSERT_HOLDING = """
INSERT INTO portfolio.etf_holding
    (etf_isin, as_of_date, constituent_isin, weight_pct, source_file)
VALUES
    (:etf_isin, :as_of_date, :constituent_isin, :weight_pct, :source_file)
ON CONFLICT (etf_isin, as_of_date, constituent_isin) DO UPDATE SET
    weight_pct  = EXCLUDED.weight_pct,
    source_file = EXCLUDED.source_file,
    loaded_at   = now()
"""


def write_to_db(constituents: list[dict], holdings: list[dict]) -> None:
    # Duplikate in constituents deduplizieren (gleiche ISIN kann in mehreren ETFs vorkommen)
    seen: set[str] = set()
    unique_constituents = []
    for c in constituents:
        if c["isin"] not in seen:
            seen.add(c["isin"])
            unique_constituents.append(c)

    db_utils.execute_many(UPSERT_CONSTITUENT, unique_constituents)
    db_utils.execute_many(UPSERT_HOLDING, holdings)


# ── Hauptprogramm ─────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Xtrackers Holdings-Parser")
    parser.add_argument("--dry-run", action="store_true",
                        help="Nur parsen, nichts in DB schreiben")
    args = parser.parse_args()

    files = sorted(HOLDINGS_DIR.glob("Constituent_*.xlsx"))
    if not files:
        log.error("Keine Constituent_*.xlsx in %s gefunden.", HOLDINGS_DIR)
        return

    log.info("Gefunden: %d Dateien", len(files))

    all_constituents: list[dict] = []
    all_holdings:     list[dict] = []

    for path in files:
        try:
            etf_isin, as_of, constituents, holdings = parse_file(path)
            log.info("%-40s  ISIN=%-14s  Datum=%s  %d Positionen",
                     path.name, etf_isin, as_of, len(holdings))
            all_constituents.extend(constituents)
            all_holdings.extend(holdings)
        except Exception as exc:
            log.error("%s – Fehler: %s", path.name, exc)

    # Gewicht-Summen prüfen
    from collections import defaultdict
    sums: dict[tuple, float] = defaultdict(float)
    for h in all_holdings:
        sums[(h["etf_isin"], h["as_of_date"])] += h["weight_pct"]
    for (isin, dt), total in sums.items():
        if abs(total - 100) > 2:
            log.warning("Gewichtssumme %s %s = %.2f %% (erwartet ~100 %%)", isin, dt, total)
        else:
            log.info("Gewichtssumme %s %s = %.2f %%", isin, dt, total)

    if args.dry_run:
        log.info("--dry-run: %d Konstituenten, %d Holdings – nichts geschrieben.",
                 len(set(c["isin"] for c in all_constituents)), len(all_holdings))
        return

    log.info("Schreibe %d unique Konstituenten → portfolio.constituent ...",
             len(set(c["isin"] for c in all_constituents)))
    log.info("Schreibe %d Holdings        → portfolio.etf_holding ...", len(all_holdings))

    try:
        write_to_db(all_constituents, all_holdings)
        log.info("Fertig.")
    except Exception as exc:
        log.exception("DB-Fehler: %s", exc)


if __name__ == "__main__":
    main()
